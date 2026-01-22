import math
import re
from collections import defaultdict

from django.core.management.base import BaseCommand
from django.db import transaction

from openpyxl import load_workbook

from apps.core.models import (
    Project,
    Dimension,
    Unit,
    ConservedEntity,
    TransformableEntity,
    Good,
    Process,
    TransformableEntityContainConservedEntity,
    GoodContainTransformableEntity,
    GoodContainGood,
    EconomicFlow,
    ElementaryFlowCompartment,
)

# ---------------------------
# Helpers
# ---------------------------


def norm(s: str) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def as_float(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        if isinstance(x, float) and (math.isnan(x) or math.isinf(x)):
            return None
        return float(x)
    try:
        x = str(x).strip()
        if x == "":
            return None
        return float(x)
    except Exception:
        return None


def iter_rows_as_dict(ws, header_row=1):
    """
    Yield dict rows based on the header row.
    Skips entirely-empty rows.
    """
    headers = []
    for cell in ws[header_row]:
        headers.append(cell.value)

    # Keep only columns that have a non-empty header
    idxs = [i for i, h in enumerate(headers) if h not in (None, "")]
    clean_headers = [headers[i] for i in idxs]

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = [row[i] for i in idxs]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        yield dict(zip(clean_headers, vals))


# ---------------------------
# Command
# ---------------------------


class Command(BaseCommand):
    help = "Import Marcot XLSX template into DB"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to .xlsx file")
        parser.add_argument(
            "--dry-run", action="store_true", help="Parse without writing"
        )
        parser.add_argument("--verbose", action="store_true", help="Verbose output")

    @transaction.atomic
    def handle(self, *args, **opts):
        path = opts["file"]
        dry = opts["dry_run"]
        verbose = opts["verbose"]

        wb = load_workbook(path, data_only=True)

        # --- caches ---
        dim_by_name = {}
        unit_by_symbol = {}
        conserved_by_extid = {}  # maps consId -> ConservedEntity
        trans_by_extid = {}  # maps id -> TransformableEntity
        good_by_extid = {}  # maps productId -> Good
        proc_by_actid = {}  # maps activityId -> Process

        # Also build activity-name matching for the "tr" sheet (which uses labels)
        proc_by_norm_name = {}

        def get_dimension(project, name):
            key = norm(name)
            if not key:
                key = "unknown"
                name = "unknown"
            if key in dim_by_name:
                return dim_by_name[key]

            obj, created = Dimension.objects.get_or_create(
                project=project,
                name=str(name).strip(),
            )

            if verbose and created:
                print("  -| Creating Dimension:", name)

            dim_by_name[key] = obj
            return obj

        def get_unit(project, symbol, dimension_name=None):
            sym = str(symbol).strip() if symbol else ""
            if not sym:
                sym = "1"  # fallback symbol for dimensionless-ish
            key = norm(sym)
            if key in unit_by_symbol:
                return unit_by_symbol[key]

            dim = get_dimension(project, dimension_name or "unknown")

            # Use symbol as name if not provided elsewhere
            obj, created = Unit.objects.get_or_create(
                symbol=sym,
                defaults={"name": sym, "dimension": dim},
            )

            if verbose and created:
                print("  -| Creating Unit:", sym, "| Dimension:", dim.name)

            # Ensure dimension is set (in case it existed but wrong/empty)
            if obj.dimension_id != dim.id:
                obj.dimension = dim
                obj.save(update_fields=["dimension"])

            unit_by_symbol[key] = obj
            return obj

        # -------------------------
        # Prepare new project
        # -------------------------
        print("Creating new project...")
        project = Project.objects.create(name="New Imported Project")

        # -------------------------
        # l_cons: conserved entities + units
        # -------------------------
        if "l_cons" in wb.sheetnames:
            print("Importing conserved entities...")

            ws = wb["l_cons"]
            for r in iter_rows_as_dict(ws):
                extid = r.get("consId")
                name = r.get("name")
                dim_name = r.get("dimension")
                unit_sym = r.get("unit")
                molar_mass = r.get("Molar mass [g/mol]")

                if not extid or not name:
                    continue

                if verbose:
                    print(
                        "-| Detected ConservedEntity:",
                        name,
                        "(ID:",
                        extid,
                        ")",
                        "| molar mass:",
                        molar_mass,
                        "g/mol",
                    )

                get_unit(project, unit_sym, dim_name)
                obj = ConservedEntity.objects.create(
                    project=project,
                    name=str(name).strip(),
                    short_name=str(extid).strip(),
                    molar_mass=molar_mass,
                )
                conserved_by_extid[str(extid).strip()] = obj

        # -------------------------
        # l_trans: transformable entities + units
        # -------------------------
        if "l_trans" in wb.sheetnames:
            ws = wb["l_trans"]
            for r in iter_rows_as_dict(ws):
                extid = r.get("id")
                name = r.get("name")
                dim_name = r.get("reference_property")
                unit_sym = r.get("unit")

                if not extid or not name:
                    continue

                if verbose:
                    print("-| Detected TransformableEntity:", name, "(ID:", extid, ")")

                get_unit(project, unit_sym, dim_name)
                obj = TransformableEntity.objects.create(
                    project=project,
                    name=str(name).strip(),
                    short_name=str(extid).strip(),
                )
                trans_by_extid[str(extid).strip()] = obj

        # -------------------------
        # l_goods: goods + reference_unit
        # -------------------------
        if "l_goods" in wb.sheetnames:
            ws = wb["l_goods"]
            for r in iter_rows_as_dict(ws):
                pid = r.get("productId")
                name = r.get("name")
                dim_name = r.get("reference_property")
                unit_sym = r.get("unit")
                if not pid or not name:
                    continue

                if verbose:
                    print("-| Detected Good:", name, "(ID:", pid, ")")

                u = get_unit(project, unit_sym, dim_name)
                obj = Good.objects.create(
                    project=project,
                    name=str(name).strip(),
                    reference_unit=u,
                )

                good_by_extid[str(pid).strip()] = obj

        # -------------------------
        # l_act: processes
        # -------------------------
        if "l_act" in wb.sheetnames:
            ws = wb["l_act"]
            for r in iter_rows_as_dict(ws):
                actid = r.get("activityId")
                name = r.get("name")
                location = r.get("location")  # TODO: Add location handling

                if not actid or not name:
                    continue

                if verbose:
                    print("-| Detected Process:", name, "(ID:", actid, ")")

                obj = Process.objects.create(project=project, name=str(name).strip())
                proc_by_actid[str(actid).strip()] = obj
                proc_by_norm_name[norm(name)] = obj

        # -------------------------
        # cons_permol: composition Transformable -> Conserved
        # -------------------------
        if "cons_permol" in wb.sheetnames:
            ws = wb["cons_permol"]
            mol_unit = get_unit(project, "mol", "amount")

            # We map substance -> TransformableEntity by extid (preferred), otherwise by name
            trans_by_name = {norm(t.name): t for t in TransformableEntity.objects.all()}
            cons_by_name = {norm(c.name): c for c in ConservedEntity.objects.all()}

            for r in iter_rows_as_dict(ws):
                substance = r.get("substance")
                element = r.get("element")
                comp = as_float(r.get("molar composition"))
                if not substance:
                    raise Exception(
                        f"Skipping cons_permol row due to missing substance: cons_permol -> {r}"
                    )
                if not element:
                    raise Exception(
                        f"Skipping cons_permol row due to missing element: cons_permol -> {r}"
                    )
                if comp is None:
                    raise Exception(
                        f"Skipping cons_permol row due to missing composition: cons_permol -> {r}"
                    )

                t = trans_by_extid.get(str(substance).strip())
                c = conserved_by_extid.get(str(element).strip())

                if not t:
                    raise Exception(
                        f"Skipping cons_permol row due to not defined transformable entity: cons_permol -> {r}"
                    )
                if not c:
                    raise Exception(
                        f"Skipping cons_permol row due to not defined conserved entity: cons_permol -> {r}"
                    )

                if verbose:
                    print(
                        "-| Linking TransformableEntity '{}' with ConservedEntity '{}' | ratio: {}".format(
                            t.name, c.name, comp
                        )
                    )

                TransformableEntityContainConservedEntity.objects.create(
                    transformable_entity=t,
                    conserved_entity=c,
                    unit=mol_unit,
                    ratio=comp,
                )

        # # -------------------------
        # # cons: composition Transformable -> Conserved
        # # -------------------------
        # if "cons" in wb.sheetnames:
        #     ws = wb["cons"]
        #     rows = list(ws.values)

        #     if not rows:
        #         pass

        #     else:
        #         header = list(rows[0])
        #         # header[0] is like "id", header[1:] are transformableIds
        #         transformable_ids = [h for h in header[1:] if h not in (None, "")]
        #         transformable_map = trans_by_extid.copy()

        #         # Build transformable lookup
        #         conservable_by_name = {
        #             norm(c.name): c for c in ConservedEntity.objects.all()
        #         }

        #         for row in rows[1:]:
        #             if not row or all(v is None for v in row):
        #                 continue

        #             conservable_key = row[0]
        #             if not conservable_key:
        #                 raise Exception(
        #                     f"Skipping cons row due to missing conserved entity: cons -> {row}"
        #                 )
        #             conservable = conserved_by_extid.get(str(conservable_key).strip()) or conservable_by_name.get(
        #                 norm(conservable_key)
        #             )
        #             if not conservable:
        #                 raise Exception(
        #                     f"Skipping cons row due to not defined conservable entity: cons -> {row}"
        #                 )

        #             for col_idx, pid in enumerate(transformable_ids, start=1):
        #                 qty = as_float(row[col_idx] if col_idx < len(row) else None)
        #                 if qty is None:
        #                     # Quantity is None, so skip-it as is 0 / not present
        #                     continue

        #                 transformable = transformable_map.get(str(pid).strip())
        #                 if not transformable:
        #                     raise Exception(
        #                         f"Skipping cons row due to not defined transformable: cons -> {row}"
        #                     )

        #                 # We want to report all the measure to 'mol' unit ?!
        #                 molar_mass = conservable.molar_mass
        #                 todo: find a way to store in Kg or to convert to mol

        #                 u = transformable.reference_unit

        #                 if verbose:
        #                     print(
        #                         "-| Linking TransformableEntity '{}' with ConservedEntity '{}' | quantity: {}".format(
        #                             transformable.name, conservable.name, qty
        #                         )
        #                     )

        #                 TransformableEntityContainConservedEntity.objects.create(
        #                     transformable_entity=transformable,
        #                     conserved_entity=conservable,
        #                     unit=mol_unit,
        #                     ratio=comp,
        #                 )

        # -------------------------
        # lay_trans: Good contains Transformable
        #    matrix: first column is transformable id/name, next columns are productIds
        # -------------------------
        if "lay_trans" in wb.sheetnames:
            ws = wb["lay_trans"]
            rows = list(ws.values)
            if not rows:
                pass
            else:
                header = list(rows[0])
                # header[0] is like "id good", header[1:] are productIds
                product_ids = [h for h in header[1:] if h not in (None, "")]
                good_map = good_by_extid.copy()

                # Build transformable lookup
                trans_by_name = {
                    norm(t.name): t for t in TransformableEntity.objects.all()
                }

                for row in rows[1:]:
                    if not row or all(v is None for v in row):
                        continue
                    trans_key = row[0]
                    if not trans_key:
                        raise Exception(
                            f"Skipping lay_trans row due to missing transformable entity: lay_trans -> {row}"
                        )
                    t = trans_by_extid.get(str(trans_key).strip()) or trans_by_name.get(
                        norm(trans_key)
                    )
                    if not t:
                        raise Exception(
                            f"Skipping lay_trans row due to not defined transformable entity: lay_trans -> {row}"
                        )

                    for col_idx, pid in enumerate(product_ids, start=1):
                        qty = as_float(row[col_idx] if col_idx < len(row) else None)
                        if qty is None:
                            # Quantity is None, so skip-it as is 0 / not present
                            continue

                        g = good_map.get(str(pid).strip())
                        if not g:
                            raise Exception(
                                f"Skipping lay_trans row due to not defined good: lay_trans -> {row}"
                            )

                        # Unit choice: best available in your file is the transformable reference unit,
                        # but you don't store it on TransformableEntity. We'll fallback to good.reference_unit.
                        # (If you later add a reference_unit on TransformableEntity, swap this.)
                        u = g.reference_unit

                        if verbose:
                            print(
                                "-| Linking Good '{}' with TransformableEntity '{}' | quantity: {}".format(
                                    g.name, t.name, qty
                                )
                            )

                        GoodContainTransformableEntity.objects.create(
                            good=g,
                            transformable_entity=t,
                            unit=u,
                            quantity=qty,
                        )

        # -------------------------
        # lay_goods: parent good contains child good
        # -------------------------
        if "lay_goods" in wb.sheetnames:
            ws = wb["lay_goods"]
            for r in iter_rows_as_dict(ws):
                parent_pid = r.get("productId")
                child_pid = r.get("ID of products inside product")
                qty = as_float(r.get("value"))

                if not parent_pid:
                    raise Exception(
                        f"Skipping lay_goods row due to missing parent productId: lay_goods -> {r}"
                    )
                if not child_pid:
                    raise Exception(
                        f"Skipping lay_goods row due to missing child productId: lay_goods -> {r}"
                    )

                if qty is None:
                    # Quantity is None, so skip-it as is 0 / not present
                    continue

                parent = good_by_extid.get(str(parent_pid).strip())
                child = good_by_extid.get(str(child_pid).strip())
                if not parent:
                    raise Exception(
                        f"Skipping lay_goods row due to not defined parent good: lay_goods -> {r}"
                    )
                if not child:
                    raise Exception(
                        f"Skipping lay_goods row due to not defined child good: lay_goods -> {r}"
                    )

                u = child.reference_unit

                if verbose:
                    print(
                        "-| Linking Good '{}' with sub-Good '{}' | quantity: {}".format(
                            parent.name, child.name, qty
                        )
                    )

                GoodContainGood.objects.create(
                    parent_good=parent,
                    child_good=child,
                    unit=u,
                    quantity=qty,
                )

        # -------------------------
        # tr: economic flows
        #     columns we use:
        #     - goods_in ID (input good)
        #     - act (activity label)  -> match Process by normalized name prefix
        #     - out ID (output good)
        #     - value (quantity)
        # -------------------------
        if "tr" in wb.sheetnames:
            ws = wb["tr"]
            procs = list(Process.objects.all())
            proc_candidates = [(norm(p.name), p) for p in procs]

            def find_process(act_label):
                k = norm(act_label)
                if not k:
                    return None
                # exact
                if k in proc_by_norm_name:
                    return proc_by_norm_name[k]
                # prefix-ish match
                for pname, pobj in proc_candidates:
                    if pname and (k.startswith(pname) or pname.startswith(k)):
                        return pobj
                return None

            for r in iter_rows_as_dict(ws):
                in_pid = r.get("goods_in ID")
                act_label = r.get("act")
                out_pid = r.get("out ID")
                qty = as_float(r.get("value"))

                if not act_label:
                    raise Exception(
                        f"Skipping tr row due to missing activity label: tr -> {r}"
                    )

                if qty is None:
                    # Quantity is None, so skip-it as is 0 / not present
                    continue

                p = find_process(act_label)

                if not p:
                    raise Exception(
                        f"Skipping tr row due to not defined Process for activity label '{act_label}': tr -> {r}"
                    )

                if in_pid:
                    g_in = good_by_extid.get(str(in_pid).strip())

                    if not g_in:
                        raise Exception(
                            f"Skipping tr row due to not defined input Good: tr -> {r}"
                        )

                    if verbose:
                        print(
                            "-| Creating EconomicFlow for Process '{}' | input Good '{}' | quantity: {}".format(
                                p.name, g_in.name, qty
                            )
                        )

                    EconomicFlow.objects.create(
                        process=p,
                        good=g_in,
                        unit=g_in.reference_unit,
                        direction="input",
                        quantity=qty,
                        is_byproduct=False,
                    )

                if out_pid:
                    g_out = good_by_extid.get(str(out_pid).strip())

                    if not g_out:
                        raise Exception(
                            f"Skipping tr row due to not defined output Good: tr -> {r}"
                        )

                    if verbose:
                        print(
                            "-| Creating EconomicFlow for Process '{}' | output Good '{}' | quantity: {}".format(
                                p.name, g_out.name, qty
                            )
                        )

                    EconomicFlow.objects.create(
                        process=p,
                        good=g_out,
                        unit=g_out.reference_unit,
                        direction="output",
                        quantity=qty,
                        is_byproduct=False,
                    )

        # -------------------------
        # background_biosphere: create compartment hierarchy
        # -------------------------
        if "background_biosphere" in wb.sheetnames:
            ws = wb["background_biosphere"]

            # Cache by (name,parent_id) to avoid duplicates
            comp_cache = {}

            def get_compartment(name, parent=None):
                key = (str(name).strip(), parent.id if parent else None)
                if key in comp_cache:
                    return comp_cache[key]
                obj, created = ElementaryFlowCompartment.objects.get_or_create(
                    project=project,
                    name=str(name).strip(),
                    parent_compartment=parent,
                )

                if verbose and created:
                    print(
                        "  -| Creating ElementaryFlowCompartment:",
                        name,
                        "| Parent:",
                        parent.name if parent else "None",
                    )

                comp_cache[key] = obj
                return obj

            for r in iter_rows_as_dict(ws):
                comp = r.get("comp")
                sub = r.get("subcomp")

                if not comp:
                    raise Exception(
                        f"Skipping background_biosphere row due to missing compartment name: background_biosphere -> {r}"
                    )

                parent = get_compartment(comp, None)
                if sub:
                    get_compartment(sub, parent)

        if dry:
            # If dry-run, rollback the whole transaction
            print("Dry-run complete (rolling back).")
            transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS("Import complete."))
